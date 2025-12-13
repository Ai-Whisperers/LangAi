"""
Excel Report Exporter (Phase 16.2).

Excel workbook generation with:
- Multiple worksheets for different data
- Formatted tables and metrics
- Charts and visualizations
- Data validation
- Professional styling

Supports both openpyxl and xlsxwriter with fallback.
"""

from typing import Dict, Any, List, Optional
from dataclasses import dataclass, field
from enum import Enum
import os
from ..utils import utc_now

# Try to import Excel libraries
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False


# ============================================================================
# Data Models
# ============================================================================

class SheetType(str, Enum):
    """Types of worksheets."""
    SUMMARY = "summary"
    FINANCIAL = "financial"
    MARKET = "market"
    COMPETITIVE = "competitive"
    METRICS = "metrics"
    RAW_DATA = "raw_data"


@dataclass
class ExcelConfig:
    """Configuration for Excel export."""
    company_name: str
    include_charts: bool = True
    include_raw_data: bool = False
    sheet_types: List[SheetType] = field(default_factory=lambda: [
        SheetType.SUMMARY,
        SheetType.FINANCIAL,
        SheetType.MARKET,
        SheetType.COMPETITIVE,
        SheetType.METRICS
    ])
    date: str = field(default_factory=lambda: utc_now().strftime("%Y-%m-%d"))


# ============================================================================
# Excel Exporter
# ============================================================================

class ExcelExporter:
    """
    Export research data to Excel workbook.

    Usage:
        exporter = ExcelExporter()
        path = exporter.export(
            research_data=data,
            config=ExcelConfig(company_name="Tesla")
        )
    """

    # Style definitions
    COLORS = {
        "header_bg": "2C5282",
        "header_fg": "FFFFFF",
        "alt_row": "EDF2F7",
        "border": "CBD5E0",
        "accent": "2B6CB0",
        "success": "38A169",
        "warning": "D69E2E",
        "danger": "E53E3E"
    }

    def __init__(self):
        """Initialize exporter."""
        self._use_openpyxl = OPENPYXL_AVAILABLE
        self._styles_created = False

    def export(
        self,
        research_data: Dict[str, Any],
        config: ExcelConfig,
        output_path: Optional[str] = None
    ) -> str:
        """
        Export research data to Excel.

        Args:
            research_data: Research data from agents
            config: Export configuration
            output_path: Output file path

        Returns:
            Path to generated Excel file
        """
        # Generate output path
        if not output_path:
            safe_name = config.company_name.lower().replace(" ", "_")
            output_path = f"./reports/{safe_name}_data_{config.date}.xlsx"

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        if self._use_openpyxl:
            return self._export_openpyxl(research_data, config, output_path)
        elif XLSXWRITER_AVAILABLE:
            return self._export_xlsxwriter(research_data, config, output_path)
        else:
            return self._export_csv_fallback(research_data, config, output_path)

    def _export_openpyxl(
        self,
        data: Dict[str, Any],
        config: ExcelConfig,
        output_path: str
    ) -> str:
        """Export using openpyxl."""
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create styles
        header_style = self._create_header_style()
        data_style = self._create_data_style()

        agent_outputs = data.get("agent_outputs", data)

        # Create sheets based on config
        for sheet_type in config.sheet_types:
            if sheet_type == SheetType.SUMMARY:
                self._create_summary_sheet(wb, agent_outputs, config, header_style, data_style)
            elif sheet_type == SheetType.FINANCIAL:
                self._create_financial_sheet(wb, agent_outputs, header_style, data_style)
            elif sheet_type == SheetType.MARKET:
                self._create_market_sheet(wb, agent_outputs, header_style, data_style)
            elif sheet_type == SheetType.COMPETITIVE:
                self._create_competitive_sheet(wb, agent_outputs, header_style, data_style)
            elif sheet_type == SheetType.METRICS:
                self._create_metrics_sheet(wb, agent_outputs, config, header_style, data_style)
            elif sheet_type == SheetType.RAW_DATA:
                self._create_raw_data_sheet(wb, agent_outputs, header_style, data_style)

        # Save workbook
        wb.save(output_path)
        return output_path

    def _create_header_style(self) -> NamedStyle:
        """Create header cell style."""
        style = NamedStyle(name="header_style")
        style.font = Font(bold=True, color=self.COLORS["header_fg"])
        style.fill = PatternFill(
            start_color=self.COLORS["header_bg"],
            end_color=self.COLORS["header_bg"],
            fill_type="solid"
        )
        style.alignment = Alignment(horizontal="center", vertical="center")
        style.border = Border(
            bottom=Side(style="thin", color=self.COLORS["border"])
        )
        return style

    def _create_data_style(self) -> NamedStyle:
        """Create data cell style."""
        style = NamedStyle(name="data_style")
        style.alignment = Alignment(horizontal="left", vertical="center")
        style.border = Border(
            bottom=Side(style="thin", color=self.COLORS["border"])
        )
        return style

    def _create_summary_sheet(
        self,
        wb,
        data: Dict[str, Any],
        config: ExcelConfig,
        header_style,
        data_style
    ):
        """Create summary worksheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"{config.company_name} Research Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Date
        ws["A2"] = f"Generated: {config.date}"
        ws["A2"].font = Font(italic=True, color="666666")

        # Key metrics section
        row = 4
        ws[f"A{row}"] = "Key Metrics"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        row += 1
        headers = ["Metric", "Value", "Status", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.COLORS["header_fg"])
            cell.fill = PatternFill(
                start_color=self.COLORS["header_bg"],
                end_color=self.COLORS["header_bg"],
                fill_type="solid"
            )

        # Extract and add metrics
        metrics = self._extract_all_metrics(data)
        for metric_name, metric_data in metrics.items():
            row += 1
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=2, value=str(metric_data.get("value", "N/A")))
            ws.cell(row=row, column=3, value=metric_data.get("status", ""))
            ws.cell(row=row, column=4, value=metric_data.get("notes", ""))

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40

    def _create_financial_sheet(self, wb, data: Dict[str, Any], header_style, data_style):
        """Create financial data worksheet."""
        ws = wb.create_sheet("Financial")

        financial_data = data.get("financial", {})

        # Title
        ws["A1"] = "Financial Analysis"
        ws["A1"].font = Font(bold=True, size=14)

        # Headers
        row = 3
        headers = ["Category", "Metric", "Value", "YoY Change", "Assessment"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.COLORS["header_fg"])
            cell.fill = PatternFill(
                start_color=self.COLORS["header_bg"],
                end_color=self.COLORS["header_bg"],
                fill_type="solid"
            )

        # Add financial data
        row += 1
        financial_metrics = [
            ("Revenue", "Total Revenue", financial_data.get("revenue", "N/A"), "", ""),
            ("Revenue", "Revenue Growth", financial_data.get("growth_rate", "N/A"), "", ""),
            ("Profitability", "Gross Margin", financial_data.get("gross_margin", "N/A"), "", ""),
            ("Profitability", "Operating Margin", financial_data.get("operating_margin", "N/A"), "", ""),
            ("Profitability", "Net Margin", financial_data.get("net_margin", "N/A"), "", ""),
            ("Health", "Debt-to-Equity", financial_data.get("debt_to_equity", "N/A"), "", ""),
            ("Health", "Current Ratio", financial_data.get("current_ratio", "N/A"), "", ""),
        ]

        for metric in financial_metrics:
            for col, value in enumerate(metric, 1):
                ws.cell(row=row, column=col, value=str(value))
            row += 1

        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_market_sheet(self, wb, data: Dict[str, Any], header_style, data_style):
        """Create market analysis worksheet."""
        ws = wb.create_sheet("Market")

        market_data = data.get("market", {})

        ws["A1"] = "Market Analysis"
        ws["A1"].font = Font(bold=True, size=14)

        # TAM/SAM/SOM section
        row = 3
        ws[f"A{row}"] = "Market Sizing"
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        headers = ["Segment", "Size ($B)", "Growth Rate", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.COLORS["header_fg"])
            cell.fill = PatternFill(
                start_color=self.COLORS["header_bg"],
                end_color=self.COLORS["header_bg"],
                fill_type="solid"
            )

        market_segments = [
            ("TAM", market_data.get("tam", "N/A"), market_data.get("tam_growth", ""), "Total Addressable Market"),
            ("SAM", market_data.get("sam", "N/A"), market_data.get("sam_growth", ""), "Serviceable Addressable Market"),
            ("SOM", market_data.get("som", "N/A"), market_data.get("som_growth", ""), "Serviceable Obtainable Market"),
        ]

        for segment in market_segments:
            row += 1
            for col, value in enumerate(segment, 1):
                ws.cell(row=row, column=col, value=str(value))

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_competitive_sheet(self, wb, data: Dict[str, Any], header_style, data_style):
        """Create competitive analysis worksheet."""
        ws = wb.create_sheet("Competitive")

        competitor_data = data.get("competitor", {})

        ws["A1"] = "Competitive Landscape"
        ws["A1"].font = Font(bold=True, size=14)

        row = 3
        headers = ["Competitor", "Type", "Threat Level", "Market Share", "Key Strength"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.COLORS["header_fg"])
            cell.fill = PatternFill(
                start_color=self.COLORS["header_bg"],
                end_color=self.COLORS["header_bg"],
                fill_type="solid"
            )

        # Add competitor data if available
        competitors_found = competitor_data.get("competitors_found", 0)
        if competitors_found > 0:
            row += 1
            ws.cell(row=row, column=1, value=f"Competitors identified: {competitors_found}")

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_metrics_sheet(
        self,
        wb,
        data: Dict[str, Any],
        config: ExcelConfig,
        header_style,
        data_style
    ):
        """Create metrics dashboard worksheet."""
        ws = wb.create_sheet("Metrics Dashboard")

        ws["A1"] = f"{config.company_name} Key Performance Metrics"
        ws["A1"].font = Font(bold=True, size=16)

        # Extract all metrics
        metrics = self._extract_all_metrics(data)

        row = 3
        headers = ["Category", "Metric", "Value", "Benchmark", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.COLORS["header_fg"])
            cell.fill = PatternFill(
                start_color=self.COLORS["header_bg"],
                end_color=self.COLORS["header_bg"],
                fill_type="solid"
            )

        for metric_name, metric_data in metrics.items():
            row += 1
            ws.cell(row=row, column=1, value=metric_data.get("category", "General"))
            ws.cell(row=row, column=2, value=metric_name)
            ws.cell(row=row, column=3, value=str(metric_data.get("value", "N/A")))
            ws.cell(row=row, column=4, value=str(metric_data.get("benchmark", "")))
            ws.cell(row=row, column=5, value=metric_data.get("status", ""))

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_raw_data_sheet(self, wb, data: Dict[str, Any], header_style, data_style):
        """Create raw data worksheet."""
        ws = wb.create_sheet("Raw Data")

        ws["A1"] = "Raw Agent Outputs"
        ws["A1"].font = Font(bold=True, size=14)

        row = 3
        for agent_name, agent_data in data.items():
            ws.cell(row=row, column=1, value=agent_name.upper())
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            if isinstance(agent_data, dict):
                for key, value in agent_data.items():
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=str(value)[:1000])
                    row += 1
            else:
                ws.cell(row=row, column=2, value=str(agent_data)[:1000])
                row += 1

            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 100

    def _extract_all_metrics(self, data: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
        """Extract all metrics from agent outputs."""
        metrics = {}

        # Financial metrics
        if "financial" in data:
            fin = data["financial"]
            if "revenue" in fin:
                metrics["Revenue"] = {"value": fin["revenue"], "category": "Financial"}
            if "growth_rate" in fin:
                metrics["Revenue Growth"] = {"value": fin["growth_rate"], "category": "Financial"}
            if "cost" in fin:
                metrics["Analysis Cost"] = {"value": f"${fin['cost']:.4f}", "category": "Operational"}

        # Market metrics
        if "market" in data:
            mkt = data["market"]
            if "tam" in mkt:
                metrics["TAM"] = {"value": mkt["tam"], "category": "Market"}
            if "market_share" in mkt:
                metrics["Market Share"] = {"value": mkt["market_share"], "category": "Market"}

        # Brand metrics
        if "brand" in data:
            brand = data["brand"]
            if "brand_score" in brand:
                metrics["Brand Score"] = {"value": brand["brand_score"], "category": "Brand"}
            if "overall_strength" in brand:
                metrics["Brand Strength"] = {"value": brand["overall_strength"], "category": "Brand"}

        # Investment metrics
        if "investment" in data:
            inv = data["investment"]
            if "investment_rating" in inv:
                metrics["Investment Rating"] = {"value": inv["investment_rating"], "category": "Investment"}
            if "overall_risk" in inv:
                metrics["Risk Level"] = {"value": inv["overall_risk"], "category": "Investment"}

        # Sales metrics
        if "sales" in data:
            sales = data["sales"]
            if "lead_score" in sales:
                metrics["Lead Score"] = {"value": sales["lead_score"], "category": "Sales"}

        return metrics

    def _export_xlsxwriter(
        self,
        data: Dict[str, Any],
        config: ExcelConfig,
        output_path: str
    ) -> str:
        """Export using xlsxwriter (alternative)."""
        workbook = xlsxwriter.Workbook(output_path)

        # Create formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#2C5282',
            'font_color': 'white',
            'align': 'center'
        })

        # Add summary sheet
        summary_ws = workbook.add_worksheet("Summary")
        summary_ws.write("A1", f"{config.company_name} Research Summary")
        summary_ws.write("A2", f"Generated: {config.date}")

        workbook.close()
        return output_path

    def _export_csv_fallback(
        self,
        data: Dict[str, Any],
        config: ExcelConfig,
        output_path: str
    ) -> str:
        """Export to CSV as fallback."""
        import csv

        csv_path = output_path.replace(".xlsx", ".csv")

        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([f"{config.company_name} Research Data"])
            writer.writerow([f"Generated: {config.date}"])
            writer.writerow([])

            agent_outputs = data.get("agent_outputs", data)

            for agent_name, agent_data in agent_outputs.items():
                writer.writerow([agent_name.upper()])
                if isinstance(agent_data, dict):
                    for key, value in agent_data.items():
                        writer.writerow([key, str(value)[:500]])
                writer.writerow([])

        return csv_path


# ============================================================================
# Factory Function
# ============================================================================

def export_to_excel(
    research_data: Dict[str, Any],
    company_name: str,
    output_path: Optional[str] = None,
    include_charts: bool = True
) -> str:
    """
    Export research data to Excel.

    Args:
        research_data: Research data from agents
        company_name: Company name
        output_path: Output file path
        include_charts: Whether to include charts

    Returns:
        Path to generated Excel file
    """
    config = ExcelConfig(
        company_name=company_name,
        include_charts=include_charts
    )

    exporter = ExcelExporter()
    return exporter.export(research_data, config, output_path)
