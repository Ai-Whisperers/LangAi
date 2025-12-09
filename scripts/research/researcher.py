"""
Comprehensive Researcher Module.

Main research orchestration class that coordinates:
- Query generation
- Web searches (with DuckDuckGo fallback)
- Analysis with Claude
- Report generation

Optimized with patterns from:
- gpt-researcher: Plan-Execute architecture
- tavily_company_researcher: 5-stage pipeline
- ai-langgraph-multi-agent: Max 4 queries pattern
- agentic_search_openai_langgraph: DuckDuckGo fallback
"""

import os
import asyncio
import yaml
import re
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

from anthropic import Anthropic
from dotenv import load_dotenv

from .config import (
    CompanyProfile,
    MarketConfig,
    ResearchResult,
    ResearchDepth,
)
from .search_provider import (
    OptimizedSearchClient,
    SearchConfig,
    SearchResult,
    SearchStrategy,
    generate_optimized_queries,
    create_search_client,
)
from .financial_provider import FinancialDataProvider, FinancialData
from .config_loader import load_config, get_config, ResearchConfig
from .research_memory import ResearchMemory, create_research_memory
from .reference_data import (
    GAP_QUERY_TEMPLATES,
    REPORT_SECTIONS,
    FINANCIAL_API_GAPS,
    GAP_PATTERNS,
    GAP_PRIORITY,
    COMPANY_TICKER_MAP,
    COMPANY_STATIC_DATA,
    get_ticker_for_company,
    get_static_company_data,
)

load_dotenv()


# Reference data is now imported from reference_data.py
# See that file for: GAP_QUERY_TEMPLATES, REPORT_SECTIONS, GAP_PATTERNS,
# GAP_PRIORITY, COMPANY_TICKER_MAP, COMPANY_STATIC_DATA


# Output file configurations for multi-file reports
OUTPUT_FILES = {
    "executive_summary": {
        "filename": "01_executive_summary.md",
        "sections": ["executive_summary"],
        "title": "Executive Summary",
        "description": "Quick overview of key findings",
    },
    "company_overview": {
        "filename": "02_company_overview.md",
        "sections": ["company_overview"],
        "title": "Company Overview",
        "description": "Company history, mission, and leadership",
    },
    "financials": {
        "filename": "03_financials.md",
        "sections": ["financial_performance"],
        "title": "Financial Performance",
        "description": "Revenue, profitability, and financial metrics",
    },
    "market_position": {
        "filename": "04_market_position.md",
        "sections": ["market_position", "products_services"],
        "title": "Market Position & Products",
        "description": "Market cap, valuation, products and services",
    },
    "competitive_analysis": {
        "filename": "05_competitive_analysis.md",
        "sections": ["competitive_landscape", "swot_analysis"],
        "title": "Competitive Analysis",
        "description": "Competitors, market share, and SWOT analysis",
    },
    "strategy": {
        "filename": "06_strategy.md",
        "sections": ["strategic_initiatives", "recent_developments"],
        "title": "Strategy & Recent Developments",
        "description": "Strategic initiatives and recent news",
    },
    "sources": {
        "filename": "07_sources.md",
        "sections": [],
        "title": "Sources",
        "description": "All sources used in research",
    },
    "full_report": {
        "filename": "00_full_report.md",
        "sections": ["all"],
        "title": "Full Research Report",
        "description": "Complete comprehensive report",
    },
}


class ComprehensiveResearcher:
    """
    Comprehensive company research orchestrator.

    Coordinates the full research pipeline:
    1. Query generation from profile
    2. Web search execution
    3. Analysis with Claude
    4. Report generation in multiple formats
    """

    def __init__(
        self,
        output_base: Optional[str] = None,
        formats: Optional[List[str]] = None,
        depth: Optional[str] = None,
        use_cache: Optional[bool] = None,
        cache_ttl_days: Optional[int] = None,
        force_refresh: Optional[bool] = None,
        # Iterative gap-filling settings
        max_iterations: Optional[int] = None,
        min_quality_score: Optional[float] = None,
        max_gaps_allowed: Optional[int] = None,
        fill_gaps: Optional[bool] = None,
        # FREE-FIRST STRATEGY SETTINGS
        search_strategy: Optional[str] = None,
        min_free_sources: Optional[int] = None,
        tavily_refinement: Optional[bool] = None,
        # CONFIG FILE SUPPORT
        use_config_file: bool = True,
        config: Optional[ResearchConfig] = None,
    ):
        """
        Initialize researcher.

        Settings can be provided via:
        1. Config file (research_config.yaml) - default
        2. ResearchConfig object passed directly
        3. Individual parameters (override config file)

        Args:
            output_base: Base directory for outputs (unified for all research)
            formats: Output formats (md, pdf, excel)
            depth: Research depth (quick, standard, comprehensive)
            use_cache: Whether to use search result caching
            cache_ttl_days: Cache TTL in days
            force_refresh: Force refresh of cached data
            max_iterations: Maximum gap-filling iterations
            min_quality_score: Minimum acceptable quality score
            max_gaps_allowed: Maximum gaps to tolerate before stopping
            fill_gaps: Whether to run iterative gap-filling
            search_strategy: Search strategy - "free_first", "auto", "free_only", "tavily_only"
            min_free_sources: Minimum sources to collect from free providers
            tavily_refinement: Use Tavily for refinement after free search
            use_config_file: Load settings from research_config.yaml (default: True)
            config: ResearchConfig object (overrides config file)
        """
        # Load config from file or use provided config
        if config:
            cfg = config
        elif use_config_file:
            cfg = get_config()
        else:
            cfg = ResearchConfig()

        # Apply settings from config, with parameter overrides
        self.output_base = Path(output_base or cfg.output.base_dir)
        self.formats = formats or cfg.output.formats
        self.depth = depth or cfg.depth
        self.use_cache = use_cache if use_cache is not None else cfg.cache.enabled
        self.cache_ttl_days = cache_ttl_days or cfg.cache.ttl_days
        self.force_refresh = force_refresh if force_refresh is not None else cfg.cache.force_refresh

        # Iterative gap-filling settings
        self.max_iterations = max_iterations or cfg.gap_filling.max_iterations
        self.min_quality_score = min_quality_score or cfg.gap_filling.min_quality_score
        self.max_gaps_allowed = max_gaps_allowed or cfg.gap_filling.max_gaps_allowed
        self.fill_gaps = fill_gaps if fill_gaps is not None else cfg.gap_filling.enabled

        # FREE-FIRST STRATEGY SETTINGS
        self.search_strategy = search_strategy or cfg.search.strategy
        self.min_free_sources = min_free_sources or cfg.search.min_free_sources
        self.tavily_refinement = tavily_refinement if tavily_refinement is not None else cfg.search.tavily_refinement

        # PREVIOUS RESEARCH REUSE SETTINGS
        self.reuse_previous_research = cfg.cache.reuse_previous_research
        self.verify_previous_content = cfg.cache.verify_previous_content
        self.max_previous_age_days = cfg.cache.max_previous_age_days

        # Initialize Anthropic client
        self.anthropic_client = Anthropic(
            api_key=os.getenv("ANTHROPIC_API_KEY")
        )

        # Initialize cache if enabled
        self.cache = None
        if self.use_cache:
            try:
                from src.company_researcher.cache import create_cache
                self.cache = create_cache(
                    cache_dir=self.output_base / ".cache",
                    ttl_days=self.cache_ttl_days,
                    enabled=True
                )
            except ImportError:
                print("[WARNING] Cache module not available, running without cache")

        # Initialize optimized search client with FREE-FIRST strategy
        search_config = SearchConfig(
            search_strategy=self.search_strategy,
            min_free_sources=self.min_free_sources,
            tavily_refinement=self.tavily_refinement,
        )
        self.search_client = create_search_client(
            cache=self.cache,
            config=search_config
        )

        # Initialize financial data provider (yfinance - free, no API key)
        self.financial_provider = FinancialDataProvider()
        if not self.financial_provider.is_available():
            print("[WARNING] Financial data provider not available (yfinance not installed)")

        # Initialize research memory for reusing previous research
        self.research_memory = create_research_memory(
            output_base=str(self.output_base),
            max_age_days=self.max_previous_age_days,
            min_quality_score=70.0,  # Lower threshold for reuse
            verify_content=self.verify_previous_content,
        )

        # Print status of previous research
        if self.reuse_previous_research:
            self.research_memory.print_status()

        # Track query history to avoid duplicates across iterations
        self._query_history: set = set()
        # Track whether financial API data was successfully retrieved
        self._has_financial_api_data: bool = False
        # Store financial data for report generation
        self._financial_data: Optional[FinancialData] = None

    async def research_company(
        self,
        company_name: str,
        profile: Optional[CompanyProfile] = None
    ) -> ResearchResult:
        """
        Research a single company with iterative gap-filling.

        Args:
            company_name: Company name to research
            profile: Optional company profile with additional context

        Returns:
            ResearchResult with all findings
        """
        start_time = datetime.now()
        print(f"\n{'='*70}")
        print(f"[RESEARCH] Starting research for: {company_name}")
        print(f"{'='*70}")

        # Reset tracking variables for this company
        self._query_history = set()
        self._has_financial_api_data = False
        self._financial_data = None

        try:
            # Step 0: Check for previous research (token-saving optimization)
            previous_sources = []
            if self.reuse_previous_research:
                can_reuse, prev_data = self.research_memory.get_reusable_data(company_name)
                if can_reuse:
                    print(f"\n[PREVIOUS RESEARCH] Found valid previous research:")
                    print(f"  - Sources: {prev_data.get('sources_count', 0)}")
                    print(f"  - Quality: {prev_data.get('quality_score', 0):.1f}")
                    print(f"  - Age: {prev_data.get('age_days', 0):.1f} days")

                    if prev_data.get('needs_refresh'):
                        print(f"  - Status: Needs refresh (will enhance with new sources)")
                        previous_sources = prev_data.get('sources', [])
                    else:
                        print(f"  - Status: Valid (will verify and update if needed)")
                        previous_sources = prev_data.get('sources', [])

                        # If we have high-quality recent research, ask user if they want to skip
                        if prev_data.get('quality_score', 0) >= 85.0 and prev_data.get('age_days', 0) < 7:
                            print(f"\n  [INFO] High-quality recent research exists.")
                            print(f"         Research will proceed to verify and potentially update.")

            # Step 1: Generate initial queries (more queries for free_first strategy)
            queries = self._generate_queries(company_name, profile)
            # Track initial queries
            for q in queries:
                self._query_history.add(q.lower())

            strategy_label = "FREE-FIRST" if self.search_strategy in ("free_first", "free_only") else self.search_strategy.upper()
            print(f"\n[QUERIES] Generated {len(queries)} search queries (strategy: {strategy_label})")

            # Step 2: Execute initial searches (Phase 1: Free sources)
            print("\n[SEARCH PHASE 1] Executing free web searches...")
            search_results = await self._execute_searches(queries)

            # Merge with previous sources if available (avoiding duplicates)
            if previous_sources:
                existing_urls = {s.get('url') for s in search_results.get("sources", [])}
                merged_count = 0
                for prev_source in previous_sources:
                    if prev_source.get('url') and prev_source.get('url') not in existing_urls:
                        search_results["sources"].append({
                            **prev_source,
                            "provider": "previous_research",
                        })
                        existing_urls.add(prev_source.get('url'))
                        merged_count += 1
                if merged_count > 0:
                    print(f"  [MERGED] Added {merged_count} sources from previous research")

            free_source_count = len(search_results.get("sources", []))

            # Step 2b: Tavily refinement phase (only if free_first + tavily_refinement enabled)
            if (self.search_strategy == "free_first" and
                self.tavily_refinement and
                free_source_count >= self.min_free_sources):
                print("\n[SEARCH PHASE 2] Tavily refinement for high-quality sources...")
                # Generate targeted Tavily queries for key topics
                tavily_queries = [
                    f'"{company_name}" investor relations financial report 2024',
                    f'"{company_name}" SEC filing 10-K annual report',
                    f'"{company_name}" earnings call transcript Q4 2024',
                ]
                tavily_results = await self.search_client.search_tavily_refinement(tavily_queries)
                if tavily_results:
                    # Merge Tavily results with existing
                    for item in tavily_results:
                        content_text = item.content or ""
                        snippet = content_text[:200].strip() + ("..." if len(content_text) > 200 else "")
                        search_results["sources"].append({
                            "title": item.title,
                            "url": item.url,
                            "score": item.score,
                            "query": item.query,
                            "provider": "tavily_refinement",
                            "snippet": snippet,
                        })
                        search_results["content"].append({
                            "title": item.title,
                            "content": item.content,
                        })
                    print(f"  [TAVILY REFINE] Added {len(tavily_results)} high-quality sources")
                    print(f"  [TOTAL] Now have {len(search_results['sources'])} sources")

            # Step 3: Initial analysis with Claude
            print("\n[ANALYSIS] Analyzing sources with Claude...")
            analysis = await self._analyze_company(
                company_name, profile, search_results
            )

            # Step 4: Iterative gap-filling (if enabled)
            iteration = 0
            total_gap_queries = 0

            if self.fill_gaps:
                while iteration < self.max_iterations:
                    # Detect gaps in the current analysis
                    gaps = self._detect_gaps(analysis.get("summary", ""))

                    # Check if we've reached acceptable quality
                    quality_score = analysis.get("quality_score", 0)
                    high_priority_gaps = [g for g in gaps if GAP_PRIORITY.get(g, 0) >= 7]

                    if quality_score >= self.min_quality_score and len(high_priority_gaps) <= self.max_gaps_allowed:
                        print(f"\n[QUALITY] Score {quality_score:.1f} >= {self.min_quality_score}, {len(high_priority_gaps)} high-priority gaps <= {self.max_gaps_allowed}")
                        print(f"[QUALITY] Research meets quality threshold. Stopping iterations.")
                        break

                    if not gaps:
                        print(f"\n[GAP-FILL] No gaps detected. Research complete.")
                        break

                    iteration += 1
                    print(f"\n[GAP-FILL] Iteration {iteration}/{self.max_iterations}")
                    print(f"  Detected {len(gaps)} gaps: {', '.join(gaps)}")
                    print(f"  High-priority gaps: {len(high_priority_gaps)}")

                    # Generate targeted queries for gaps
                    gap_queries = self._generate_gap_queries(
                        company_name, profile, gaps, max_queries=4
                    )

                    if not gap_queries:
                        print(f"  No new queries generated. Stopping.")
                        break

                    total_gap_queries += len(gap_queries)
                    print(f"  Generated {len(gap_queries)} targeted queries")

                    # Execute gap-filling searches
                    gap_results = await self._execute_searches(gap_queries)

                    # Merge new results with existing
                    search_results = self._merge_search_results(search_results, gap_results)

                    # Re-analyze with enriched data
                    print(f"  Re-analyzing with {len(search_results.get('sources', []))} total sources...")
                    analysis = await self._analyze_company(
                        company_name, profile, search_results
                    )

                if iteration > 0:
                    print(f"\n[GAP-FILL] Completed {iteration} iteration(s), {total_gap_queries} additional queries")

            # Step 5: Generate final reports
            print("\n[REPORTS] Generating reports...")
            report_paths = await self._generate_reports(
                company_name, profile, analysis, search_results
            )

            # Calculate final metrics
            duration = (datetime.now() - start_time).total_seconds()
            cost = self._estimate_cost(analysis)
            total_queries = len(queries) + total_gap_queries

            # Final gap analysis for reporting
            final_gaps = self._detect_gaps(analysis.get("summary", ""))

            result = ResearchResult(
                company_name=company_name,
                success=True,
                profile=profile,
                summary=analysis.get("summary", ""),
                analysis=analysis,
                sources=search_results.get("sources", []),
                report_paths=report_paths,
                metrics={
                    "duration_seconds": duration,
                    "cost_usd": cost,
                    "sources_count": len(search_results.get("sources", [])),
                    "queries_count": total_queries,
                    "gap_fill_iterations": iteration,
                    "remaining_gaps": len(final_gaps),
                    "gap_categories": final_gaps
                },
                quality_score=analysis.get("quality_score", 70.0)
            )

            self._print_result_summary(result)
            return result

        except Exception as e:
            duration = (datetime.now() - start_time).total_seconds()
            print(f"\n[ERROR] Research failed: {e}")
            import traceback
            traceback.print_exc()
            return ResearchResult(
                company_name=company_name,
                success=False,
                profile=profile,
                error=str(e),
                metrics={"duration_seconds": duration}
            )

    async def research_market(
        self,
        market_folder: str,
        generate_comparison: bool = True
    ) -> List[ResearchResult]:
        """
        Research all companies in a market folder.

        Args:
            market_folder: Path to market folder with YAML files
            generate_comparison: Whether to generate comparison report

        Returns:
            List of research results
        """
        market_path = Path(market_folder)
        results = []

        # Load market config if exists
        market_config = None
        market_file = market_path / "_market.yaml"
        if market_file.exists():
            with open(market_file) as f:
                market_config = MarketConfig.from_yaml(yaml.safe_load(f))
            print(f"\n[MARKET] Loading market: {market_config.name}")
            print(f"  Country: {market_config.country}")
            print(f"  Industry: {market_config.industry}")

        # Find all company YAML files
        company_files = [
            f for f in market_path.glob("*.yaml")
            if not f.name.startswith("_")
        ]

        print(f"\n[BATCH] Found {len(company_files)} companies to research\n")

        # Research each company
        for i, yaml_file in enumerate(company_files, 1):
            print(f"\n[{i}/{len(company_files)}] Processing: {yaml_file.name}")

            with open(yaml_file) as f:
                data = yaml.safe_load(f)

            profile = CompanyProfile.from_yaml(data)
            result = await self.research_company(profile.name, profile)
            results.append(result)

        # Generate comparison report if requested
        if generate_comparison and len(results) > 1:
            await self._generate_comparison_report(results, market_config)

        # Print batch summary
        self._print_batch_summary(results)

        return results

    def _generate_queries(
        self,
        company_name: str,
        profile: Optional[CompanyProfile]
    ) -> List[str]:
        """
        Generate optimized search queries for a company.

        For free_first strategy:
        - Generates 3x more queries to reach 100+ sources
        - Uses extended query templates for comprehensive coverage

        For auto/tavily strategy:
        - Reduces query count by 33% (6 -> 4 base queries)
        - Combines related topics into single queries
        - Maximizes information per Tavily credit
        """
        return generate_optimized_queries(
            company_name=company_name,
            profile=profile,
            depth=self.depth,
            strategy=self.search_strategy
        )

    async def _execute_searches(self, queries: List[str]) -> Dict[str, Any]:
        """
        Execute web searches using the optimized search client.

        For FREE-FIRST strategy:
        - Phase 1: Use DuckDuckGo to collect 100+ free sources
        - Phase 2: (Optional) Use Tavily for high-priority gap refinement

        For AUTO strategy:
        - Automatic fallback: Tavily -> Serper -> Brave -> DuckDuckGo
        - Domain filtering for quality sources
        - Integrated caching
        """
        all_sources = []
        all_content = []

        # Show strategy being used
        if self.search_strategy in ("free_first", "free_only"):
            print(f"  [STRATEGY] FREE-FIRST: Collecting {self.min_free_sources}+ sources from DuckDuckGo")

        for query in queries:
            try:
                # Use optimized search client (handles cache, fallback, etc.)
                results = await self.search_client.search(
                    query=query,
                    use_cache=not self.force_refresh
                )

                # Process results (SearchResult objects)
                for item in results:
                    # Create a snippet from the content (first 200 chars)
                    content_text = item.content or ""
                    snippet = content_text[:200].strip()
                    if len(content_text) > 200:
                        snippet += "..."

                    all_sources.append({
                        "title": item.title,
                        "url": item.url,
                        "score": item.score,
                        "query": query,
                        "provider": item.provider,
                        "snippet": snippet,  # Add snippet for display in sources file
                    })
                    all_content.append({
                        "title": item.title,
                        "content": item.content,
                    })

            except Exception as e:
                print(f"  [ERROR] Search failed for '{query[:40]}...': {e}")
                continue

        # Print search statistics
        stats = self.search_client.get_stats()
        free_sources = stats.get('free_sources_total', 0)
        tavily_sources = stats.get('tavily_sources_total', 0)
        duckduckgo_calls = stats.get('duckduckgo_calls', 0)
        duckduckgo_successes = stats.get('duckduckgo_successes', 0)

        print(f"\n  [SEARCH RESULTS] Found {len(all_sources)} total sources")

        if self.search_strategy in ("free_first", "free_only"):
            print(f"       DuckDuckGo: {duckduckgo_successes}/{duckduckgo_calls} queries successful")
            print(f"       Free sources: {free_sources}")
            if len(all_sources) >= self.min_free_sources:
                print(f"       [OK] Target of {self.min_free_sources}+ sources ACHIEVED")
            else:
                print(f"       [WARN] Below target of {self.min_free_sources} sources ({len(all_sources)} found)")
        else:
            print(f"       Tavily: {stats['tavily_successes']}/{stats['tavily_calls']} calls")
            if stats['duckduckgo_fallbacks'] > 0:
                print(f"       DuckDuckGo fallbacks: {stats['duckduckgo_fallbacks']}")

        if stats['cache_hits'] > 0:
            print(f"       Cache hits: {stats['cache_hits']}")
        print()

        return {
            "sources": all_sources,
            "content": all_content
        }

    def _detect_gaps(self, report_text: str) -> List[str]:
        """
        Detect gaps in the generated report using both pattern and section analysis.

        Uses two approaches:
        1. Pattern matching: Looks for "Data not available" type text
        2. Section analysis: Checks if sections have sufficient content

        Args:
            report_text: The generated report content

        Returns:
            List of gap category names sorted by priority
        """
        detected_gaps = []
        report_lower = report_text.lower()

        # Skip financial gaps if API data was successfully retrieved
        skip_gaps = FINANCIAL_API_GAPS if self._has_financial_api_data else set()

        # 1. Pattern-based gap detection (explicit "Data not available" mentions)
        for gap_category, patterns in GAP_PATTERNS.items():
            if gap_category in skip_gaps:
                continue
            for pattern in patterns:
                if re.search(pattern, report_lower, re.IGNORECASE):
                    if gap_category not in detected_gaps:
                        detected_gaps.append(gap_category)
                    break

        # 2. Section-based completeness analysis
        section_completeness = self._analyze_section_completeness(report_text)
        for section_name, completeness in section_completeness.items():
            if completeness["complete"]:
                continue

            # Map sections to gap categories
            section_to_gap = {
                "financial_performance": "revenue_segments",
                "market_position": "market_cap" if "market_cap" not in skip_gaps else None,
                "competitive_landscape": "competitors",
                "company_overview": "leadership",
                "strategic_initiatives": "acquisitions",
            }

            gap = section_to_gap.get(section_name)
            if gap and gap not in detected_gaps and gap not in skip_gaps:
                detected_gaps.append(gap)

        # Sort by priority (highest first)
        detected_gaps.sort(key=lambda g: GAP_PRIORITY.get(g, 0), reverse=True)
        return detected_gaps

    def _analyze_section_completeness(self, report_text: str) -> Dict[str, Dict[str, Any]]:
        """
        Analyze completeness of each report section.

        For each defined section, extracts content and checks:
        - Section presence
        - Content length vs minimum requirement
        - Presence of required elements

        Args:
            report_text: The generated report content

        Returns:
            Dict mapping section names to completeness info
        """
        results = {}

        for section_name, config in REPORT_SECTIONS.items():
            section_info = {
                "found": False,
                "length": 0,
                "min_length": config["min_length"],
                "elements_found": [],
                "elements_missing": [],
                "complete": False,
                "score": 0.0,
            }

            # Find section content
            content = self._extract_section_content(report_text, config["patterns"])
            if content:
                section_info["found"] = True
                section_info["length"] = len(content)

                # Check for required elements
                content_lower = content.lower()
                for element in config["required_elements"]:
                    if element.lower() in content_lower:
                        section_info["elements_found"].append(element)
                    else:
                        section_info["elements_missing"].append(element)

                # Calculate completeness score
                length_score = min(1.0, section_info["length"] / config["min_length"])
                elements_score = len(section_info["elements_found"]) / len(config["required_elements"]) if config["required_elements"] else 1.0
                section_info["score"] = (length_score * 0.6) + (elements_score * 0.4)

                # Mark as complete if passes thresholds
                section_info["complete"] = (
                    section_info["length"] >= config["min_length"] * 0.7 and
                    section_info["score"] >= 0.6
                )

            results[section_name] = section_info

        return results

    def _extract_section_content(self, report_text: str, patterns: List[str]) -> str:
        """
        Extract content of a specific section from the report.

        Args:
            report_text: Full report text
            patterns: Regex patterns to match section header

        Returns:
            Section content (text between this header and next header)
        """
        for pattern in patterns:
            match = re.search(pattern, report_text, re.IGNORECASE)
            if match:
                start = match.end()
                # Find next section (any ## or # header)
                next_section = re.search(r'\n##?\s+\d*\.?\s*[A-Z]', report_text[start:])
                if next_section:
                    end = start + next_section.start()
                else:
                    end = len(report_text)
                return report_text[start:end].strip()
        return ""

    def _generate_gap_queries(
        self,
        company_name: str,
        profile: Optional[CompanyProfile],
        gaps: List[str],
        max_queries: int = 4
    ) -> List[str]:
        """
        Generate targeted search queries to fill identified gaps.

        Features:
        - Uses query history to avoid duplicates across iterations
        - Falls back to alternate templates if primary already used
        - Adds industry/country context for more relevant results

        Args:
            company_name: Company name
            profile: Company profile (for ticker, industry context)
            gaps: List of gap categories to address
            max_queries: Maximum queries to generate

        Returns:
            List of NEW targeted search queries (not previously used)
        """
        queries = []
        ticker = ""
        industry = ""
        country = ""

        if profile:
            ticker = getattr(profile, 'ticker', '') or ''
            industry = getattr(profile, 'industry', '') or ''
            country = getattr(profile, 'country', '') or ''

        # Generate queries for each gap, prioritizing high-priority gaps
        for gap in gaps:
            if len(queries) >= max_queries:
                break

            templates = GAP_QUERY_TEMPLATES.get(gap, [])

            # Try each template until we find one not in history
            for template in templates:
                query = template.format(
                    company=company_name,
                    ticker=ticker if ticker else company_name
                )

                # Check if this query (or similar) was already used
                query_lower = query.lower()
                if query_lower not in self._query_history:
                    queries.append(query)
                    self._query_history.add(query_lower)
                    break
            else:
                # All templates exhausted, try with industry/country context
                if industry and gap in ["competitors", "market_share"]:
                    context_query = f'"{company_name}" {gap.replace("_", " ")} {industry} industry'
                    if context_query.lower() not in self._query_history:
                        queries.append(context_query)
                        self._query_history.add(context_query.lower())
                elif country and gap in ["employees", "leadership"]:
                    context_query = f'"{company_name}" {gap.replace("_", " ")} {country}'
                    if context_query.lower() not in self._query_history:
                        queries.append(context_query)
                        self._query_history.add(context_query.lower())

        return queries[:max_queries]

    def _merge_search_results(
        self,
        existing: Dict[str, Any],
        new_results: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Merge new search results with existing results.

        Deduplicates by URL and combines sources/content lists.

        Args:
            existing: Existing search results
            new_results: New search results to merge

        Returns:
            Merged search results dictionary
        """
        # Get existing URLs for deduplication
        existing_urls = {s.get('url', '') for s in existing.get('sources', [])}

        merged_sources = list(existing.get('sources', []))
        merged_content = list(existing.get('content', []))

        for source in new_results.get('sources', []):
            url = source.get('url', '')
            if url and url not in existing_urls:
                merged_sources.append(source)
                existing_urls.add(url)

        for content in new_results.get('content', []):
            # Simple dedup by title
            existing_titles = {c.get('title', '') for c in merged_content}
            if content.get('title', '') not in existing_titles:
                merged_content.append(content)

        return {
            "sources": merged_sources,
            "content": merged_content
        }

    def _infer_ticker(self, company_name: str) -> Optional[str]:
        """
        Infer ticker symbol from company name using the lookup table.

        Args:
            company_name: Company name to look up

        Returns:
            Ticker symbol if found, None otherwise
        """
        # Normalize company name for lookup
        name_lower = company_name.lower().strip()

        # Direct lookup
        if name_lower in COMPANY_TICKER_MAP:
            return COMPANY_TICKER_MAP[name_lower]

        # Try removing common suffixes
        for suffix in [" inc", " inc.", " corp", " corp.", " corporation", " company", " co", " ltd", " llc"]:
            name_without_suffix = name_lower.replace(suffix, "").strip()
            if name_without_suffix in COMPANY_TICKER_MAP:
                return COMPANY_TICKER_MAP[name_without_suffix]

        return None

    async def _analyze_company(
        self,
        company_name: str,
        profile: Optional[CompanyProfile],
        search_results: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Analyze company with Claude."""
        content = search_results.get("content", [])
        combined = "\n---\n".join([
            f"Source: {c['title']}\n{c['content']}"
            for c in content[:12]
        ])

        # Fetch financial data from Yahoo Finance if ticker available
        financial_context = ""
        ticker = None

        # First try to get ticker from profile
        if profile:
            ticker = getattr(profile, 'ticker', None)

        # If no ticker from profile, try to infer from company name
        if not ticker:
            ticker = self._infer_ticker(company_name)
            if ticker:
                print(f"  [TICKER] Auto-detected ticker: {ticker} for {company_name}")

        # Look up static company data for well-known companies
        static_data = {}
        if ticker and ticker in COMPANY_STATIC_DATA:
            static_data = COMPANY_STATIC_DATA[ticker]
            print(f"  [STATIC] Found static data for {ticker}: CEO={static_data.get('ceo', 'N/A')}")

        if ticker and self.financial_provider.is_available():
            print(f"  [FINANCIAL] Fetching data for {ticker}...")
            financial_data = self.financial_provider.get_financial_data(ticker)
            if financial_data:
                # Store for report generation and mark that we have API data
                self._financial_data = financial_data
                self._has_financial_api_data = True
                financial_context = f"""
{financial_data.to_context_string()}

IMPORTANT: Use the financial data above from Yahoo Finance as the authoritative source for:
- Market capitalization, stock price, 52-week range
- P/E ratio, Forward P/E, PEG ratio
- Revenue, EPS, Dividend yield
- Sector, industry, employee count
Include these EXACT values in the report. Do NOT mark them as "Data not available".
"""
                price_str = f"${financial_data.current_price:.2f}" if financial_data.current_price else "N/A"
                print(f"  [FINANCIAL] Got data: {price_str}, Market Cap: {financial_data.market_cap_formatted or 'N/A'}")
            else:
                print(f"  [FINANCIAL] No data found for {ticker}")

        # Build context from profile with all available data
        # Use static data as fallback for well-known companies
        context = ""
        if profile:
            # Get additional details if available
            details = getattr(profile, 'details', {}) or {}
            ceo = details.get('ceo', '') if isinstance(details, dict) else ''
            headquarters = details.get('headquarters', '') if isinstance(details, dict) else ''
            employees = details.get('employees', '') if isinstance(details, dict) else ''
            ticker = getattr(profile, 'ticker', '') or ''
            exchange = getattr(profile, 'exchange', '') or ''
            website = getattr(profile, 'website', '') or ''
            founded = getattr(profile, 'founded', '') or details.get('founded', '') if isinstance(details, dict) else ''

            # Use static data as fallback
            if not ceo and static_data.get('ceo'):
                ceo = static_data['ceo']
            if not headquarters and static_data.get('headquarters'):
                headquarters = static_data['headquarters']
            if not employees and static_data.get('employees'):
                employees = static_data['employees']
            if not founded and static_data.get('founded'):
                founded = static_data['founded']

            context = f"""
Company Profile (Known Information):
- Legal Name: {profile.legal_name or company_name}
- Ticker: {ticker} ({exchange}) if available
- Website: {website}
- Industry: {profile.industry or 'N/A'}
- Country: {profile.country or 'N/A'}
- Headquarters: {headquarters}
- CEO: {ceo}
- Founded: {founded}
- Employees: {employees}
- Parent Company: {profile.parent_company or 'Independent'}
- Services: {', '.join(profile.services[:5]) if profile.services else 'N/A'}
- Key Competitors: {', '.join(profile.competitors[:5]) if profile.competitors else 'N/A'}

NOTE: Use this profile data to supplement the research. If the profile says "CEO: Tim Cook", include that in the Leadership section.
"""
        elif static_data:
            # No profile but we have static data for this well-known company
            recent_products_str = ""
            if static_data.get('recent_products'):
                recent_products_str = "\n- Recent Product Launches (2024-2025):\n  " + "\n  ".join(static_data['recent_products'])

            ai_initiatives_str = ""
            if static_data.get('ai_initiatives'):
                ai_initiatives_str = f"\n- AI Initiatives: {static_data['ai_initiatives']}"

            context = f"""
Company Profile (Known Information):
- Company: {company_name}
- Ticker: {ticker or 'N/A'}
- CEO: {static_data.get('ceo', 'N/A')}
- Founded: {static_data.get('founded', 'N/A')}
- Headquarters: {static_data.get('headquarters', 'N/A')}
- Founders: {static_data.get('founders', 'N/A')}
- Employees: {static_data.get('employees', 'N/A')}{recent_products_str}{ai_initiatives_str}

IMPORTANT: This is verified company data. Include ALL of the above information in the report:
- CEO, headquarters, founded date, and employee count in the Company Overview section
- Recent Product Launches in the Products & Services section
- AI Initiatives in the Strategic Initiatives section
Do NOT mark these as "Data not available" - they are confirmed facts.
"""

        prompt = f"""Analyze the following research sources about {company_name} and create a comprehensive research report.

{context}
{financial_context}

Research Sources:
{combined}

Create a detailed report with the following sections:

1. **Executive Summary** - 4-5 sentences covering: key financial metrics, market position, strategic direction, and notable recent developments

2. **Company Overview**
   - History and founding
   - Mission/vision
   - Leadership team (CEO, key executives)
   - Headquarters and global presence
   - Employee count

3. **Financial Performance**
   - Annual revenue (current year and YoY growth)
   - Quarterly results (if available)
   - Net income/profitability
   - Earnings per share (EPS)
   - Key financial ratios
   - Revenue by segment/business unit (if applicable)

4. **Market Position**
   - Market capitalization
   - Stock performance (current price, 52-week range)
   - Market share by segment
   - Competitive positioning
   - Valuation metrics (P/E ratio, etc.)

5. **Products & Services**
   - Major product lines/business segments
   - Revenue contribution by segment
   - Key services and solutions
   - Recent product launches (2024-2025) - EXTRACT FROM SOURCES: Look for any product announcements, new versions, hardware releases, software updates, or service launches mentioned in the research sources. List specific product names with release dates.

6. **Strategic Initiatives**
   - AI and emerging technology investments
   - Cloud/digital transformation
   - Acquisitions and partnerships
   - R&D focus areas

7. **Competitive Landscape**
   - Main competitors with market share comparison
   - Competitive advantages/moats
   - Differentiation strategy

8. **Recent Developments** (2024-2025)
   - Major announcements
   - Earnings highlights
   - Strategic moves

9. **SWOT Analysis**
   - Strengths (specific to this company)
   - Weaknesses (actual challenges, not generic)
   - Opportunities (concrete market opportunities)
   - Threats (specific competitive/regulatory threats)

10. **Key Metrics Summary** (Table format with all available data)

Requirements:
- Include SPECIFIC numbers with dates (e.g., "$245B revenue in FY2024")
- Include stock price and market cap if available
- Break down revenue by business segment if available
- Cite the source URL for key claims
- If data is missing, explicitly state "Data not available" rather than omitting
- Be objective and data-driven
- Format using Markdown with proper headers"""

        response = self.anthropic_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=6000,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}]
        )

        summary = response.content[0].text

        return {
            "summary": summary,
            "tokens": {
                "input": response.usage.input_tokens,
                "output": response.usage.output_tokens
            },
            "quality_score": self._calculate_quality_score(summary, search_results)
        }

    def _calculate_quality_score(
        self,
        summary: str,
        search_results: Dict[str, Any]
    ) -> float:
        """
        Calculate research quality score based on multiple factors.

        Scoring breakdown (max 100):
        - Base score: 30
        - Source quantity: up to 15 points
        - Section completeness: up to 30 points
        - Financial data presence: up to 15 points
        - Content depth: up to 10 points
        """
        score = 30.0

        # Source quantity bonus (up to 15 points)
        sources = search_results.get("sources", [])
        source_count = len(sources)
        if source_count >= 15:
            score += 15
        elif source_count >= 10:
            score += 12
        elif source_count >= 5:
            score += 8
        elif source_count >= 3:
            score += 5

        # Section completeness bonus (up to 30 points)
        section_analysis = self._analyze_section_completeness(summary)
        complete_sections = sum(1 for s in section_analysis.values() if s["complete"])
        total_sections = len(section_analysis)
        if total_sections > 0:
            section_score = (complete_sections / total_sections) * 30
            score += section_score

        # Financial data presence (up to 15 points)
        financial_indicators = [
            (r'\$[\d,.]+\s*(?:trillion|billion|million|T|B|M)', 5),  # Dollar amounts
            (r'market cap(?:italization)?[:\s]+\$', 3),  # Market cap mention
            (r'(?:P/?E|price.to.earnings)\s*(?:ratio)?[:\s]+[\d.]+', 3),  # P/E ratio
            (r'revenue[:\s]+\$[\d,.]+', 2),  # Revenue
            (r'(?:EPS|earnings per share)[:\s]+\$[\d.]+', 2),  # EPS
        ]
        financial_score = 0
        for pattern, points in financial_indicators:
            if re.search(pattern, summary, re.IGNORECASE):
                financial_score += points
        score += min(15, financial_score)

        # Content depth bonus (up to 10 points)
        if len(summary) > 5000:
            score += 10
        elif len(summary) > 3000:
            score += 7
        elif len(summary) > 1500:
            score += 4

        return min(100.0, score)

    async def _generate_reports(
        self,
        company_name: str,
        profile: Optional[CompanyProfile],
        analysis: Dict[str, Any],
        search_results: Dict[str, Any]
    ) -> Dict[str, str]:
        """Generate reports in multiple files."""
        # Create output directory
        safe_name = company_name.lower().replace(" ", "_").replace(".", "")
        timestamp = datetime.now().strftime("%Y%m%d")
        company_dir = self.output_base / safe_name
        company_dir.mkdir(parents=True, exist_ok=True)

        report_paths = {}
        summary_content = analysis.get('summary', '')
        sources = search_results.get("sources", [])

        # Generate Markdown files
        if "md" in self.formats:
            # Extract sections from the full report
            sections = self._extract_all_sections(summary_content)

            # Generate each output file
            for file_key, file_config in OUTPUT_FILES.items():
                filename = file_config["filename"]
                file_path = company_dir / filename

                if file_key == "full_report":
                    # Full report: complete content
                    content = self._generate_full_report(
                        company_name, profile, summary_content, sources, timestamp
                    )
                elif file_key == "sources":
                    # Sources file
                    content = self._generate_sources_file(
                        company_name, sources, timestamp
                    )
                else:
                    # Section-specific file
                    section_names = file_config["sections"]
                    section_content = self._extract_sections_content(
                        summary_content, section_names, sections
                    )
                    content = self._generate_section_file(
                        company_name, file_config["title"],
                        file_config["description"], section_content, timestamp
                    )

                file_path.write_text(content, encoding="utf-8")
                report_paths[file_key] = str(file_path)
                print(f"  [FILE] {filename}")

            # Generate metrics.json (machine-readable key metrics)
            metrics_path = company_dir / "metrics.json"
            self._write_metrics_json(metrics_path, company_name, profile, analysis, search_results)
            report_paths["metrics"] = str(metrics_path)
            print(f"  [FILE] metrics.json")

            # Generate financial_data.json if available
            if self._financial_data:
                financial_path = company_dir / "financial_data.json"
                self._write_financial_json(financial_path)
                report_paths["financial_data"] = str(financial_path)
                print(f"  [FILE] financial_data.json")

            # Generate index.md for easy navigation
            index_path = company_dir / "README.md"
            self._write_index_file(index_path, company_name, report_paths, timestamp)
            report_paths["index"] = str(index_path)
            print(f"  [FILE] README.md")

        # Generate PDF if requested
        if "pdf" in self.formats:
            try:
                pdf_path = company_dir / f"{safe_name}_report_{timestamp}.pdf"
                self._write_pdf_report(pdf_path, company_name, profile, analysis)
                report_paths["pdf"] = str(pdf_path)
                print(f"  [FILE] {pdf_path.name}")
            except Exception as e:
                print(f"  [WARNING] PDF generation failed: {e}")

        # Generate Excel if requested
        if "excel" in self.formats:
            try:
                excel_path = company_dir / f"{safe_name}_report_{timestamp}.xlsx"
                self._write_excel_report(excel_path, company_name, profile, analysis)
                report_paths["excel"] = str(excel_path)
                print(f"  [FILE] {excel_path.name}")
            except Exception as e:
                print(f"  [WARNING] Excel generation failed: {e}")

        return report_paths

    def _extract_all_sections(self, content: str) -> Dict[str, str]:
        """Extract all sections from the report content."""
        sections = {}
        for section_key, section_config in REPORT_SECTIONS.items():
            for pattern in section_config["patterns"]:
                match = re.search(pattern, content, re.IGNORECASE)
                if match:
                    # Find start and end of section
                    start = match.start()
                    # Find next section header or end of content
                    next_section = re.search(r'\n##?\s+\d*\.?\s*[A-Z]', content[match.end():])
                    if next_section:
                        end = match.end() + next_section.start()
                    else:
                        end = len(content)
                    sections[section_key] = content[start:end].strip()
                    break
        return sections

    def _extract_sections_content(
        self,
        full_content: str,
        section_names: List[str],
        extracted_sections: Dict[str, str]
    ) -> str:
        """Extract specific sections from the full content."""
        parts = []
        for name in section_names:
            if name in extracted_sections:
                parts.append(extracted_sections[name])
        return "\n\n".join(parts) if parts else "No content available for this section."

    def _generate_full_report(
        self,
        company_name: str,
        profile: Optional[CompanyProfile],
        summary_content: str,
        sources: List[Dict],
        timestamp: str
    ) -> str:
        """Generate the full comprehensive report."""
        content = f"""# {company_name} Research Report

**Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M")}

---

{summary_content}

---

## Sources

"""
        for source in sources[:20]:
            content += f"- [{source.get('title', 'Source')}]({source.get('url', '')})\n"

        return content

    def _generate_section_file(
        self,
        company_name: str,
        title: str,
        description: str,
        section_content: str,
        timestamp: str
    ) -> str:
        """Generate a section-specific file."""
        return f"""# {company_name} - {title}

**Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M")}

*{description}*

---

{section_content}

---

*This is part of the {company_name} research report. See [Full Report](00_full_report.md) for complete analysis.*
"""

    def _generate_sources_file(
        self,
        company_name: str,
        sources: List[Dict],
        timestamp: str
    ) -> str:
        """Generate the sources file with snippets."""
        content = f"""# {company_name} - Research Sources

**Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M")}

*All sources used in this research report*

---

## Primary Sources

"""
        # Group sources by query for better organization
        sources_by_query = {}
        for source in sources[:30]:
            query = source.get('query', 'General')
            if query not in sources_by_query:
                sources_by_query[query] = []
            sources_by_query[query].append(source)

        source_num = 1
        for query, query_sources in sources_by_query.items():
            # Add query header
            short_query = query[:60] + "..." if len(query) > 60 else query
            content += f"### Query: *{short_query}*\n\n"

            for source in query_sources:
                title = source.get('title', 'Source')
                url = source.get('url', '')
                snippet = source.get('snippet', '')
                provider = source.get('provider', '')

                # Clean up snippet - remove extra whitespace and newlines
                if snippet:
                    snippet = ' '.join(snippet.split())
                    # Truncate if needed
                    if len(snippet) > 180:
                        snippet = snippet[:180] + "..."

                content += f"{source_num}. **[{title}]({url})**"
                if provider:
                    content += f" *({provider})*"
                content += "\n"

                if snippet:
                    content += f"   > {snippet}\n"
                content += "\n"
                source_num += 1

            content += "\n"

        content += f"""---

## Source Statistics

- **Total Sources**: {len(sources)}
- **Unique Queries**: {len(sources_by_query)}
- **Research Date**: {datetime.now().strftime("%Y-%m-%d")}

### Source Providers
"""
        # Count sources by provider
        provider_counts = {}
        for source in sources:
            provider = source.get('provider', 'unknown')
            provider_counts[provider] = provider_counts.get(provider, 0) + 1

        for provider, count in sorted(provider_counts.items(), key=lambda x: -x[1]):
            content += f"- **{provider}**: {count} sources\n"

        content += f"""
---

*This is part of the {company_name} research report. See [Full Report](00_full_report.md) for complete analysis.*
"""
        return content

    def _write_metrics_json(
        self,
        path: Path,
        company_name: str,
        profile: Optional[CompanyProfile],
        analysis: Dict[str, Any],
        search_results: Dict[str, Any]
    ):
        """Write machine-readable metrics JSON file."""
        metrics = {
            "company": company_name,
            "generated_at": datetime.now().isoformat(),
            "profile": {
                "legal_name": profile.legal_name if profile else None,
                "ticker": profile.ticker if profile else None,
                "exchange": profile.exchange if profile else None,
                "industry": profile.industry if profile else None,
                "country": profile.country if profile else None,
            } if profile else None,
            "research": {
                "quality_score": analysis.get("quality_score"),
                "depth": self.depth,
                "iteration_count": analysis.get("iteration_count", 1),
                "sources_count": len(search_results.get("sources", [])),
            },
            "financial_data_available": self._financial_data is not None,
        }

        # Extract key metrics from the summary if possible
        summary = analysis.get("summary", "")
        extracted = self._extract_key_metrics(summary)
        if extracted:
            metrics["extracted_metrics"] = extracted

        path.write_text(json.dumps(metrics, indent=2), encoding="utf-8")

    def _extract_key_metrics(self, summary: str) -> Dict[str, Any]:
        """Extract key metrics from report text using regex."""
        metrics = {}

        # Market cap patterns - require at least 1 digit before the unit
        # Matches: "$3.5 trillion", "$3.5T", "3.5 trillion", "$3,500 billion"
        # Also handles markdown bold like "**Market Cap**: $4.14 trillion"
        mcap_patterns = [
            r'\*?\*?market\s*cap(?:italization)?\*?\*?[:\s]*\$?([\d.,]+)\s*(trillion|billion|million|T|B|M)\b',
            r'market\s*cap(?:italization)?\s+(?:of\s+)?\$?([\d.,]+)\s*(trillion|billion|million|T|B|M)\b',
            r'(?:valued at|worth)\s*\$?([\d.,]+)\s*(trillion|billion|million|T|B|M)\b',
            r'\$?([\d.,]+)\s*(trillion|billion|million|T|B|M)\s*market\s*cap',
            r'\| \*?\*?Market Cap(?:italization)?\*?\*? \| \$?([\d.,]+)(T|B|M)\b',  # Table format
        ]
        for pattern in mcap_patterns:
            mcap_match = re.search(pattern, summary, re.IGNORECASE)
            if mcap_match:
                value = mcap_match.group(1).replace(",", "")
                # Ensure we have a valid number
                try:
                    float(value)
                    unit = mcap_match.group(2).upper()
                    # Normalize unit
                    unit_map = {"TRILLION": "T", "BILLION": "B", "MILLION": "M"}
                    unit = unit_map.get(unit, unit)
                    metrics["market_cap_text"] = f"${value}{unit}"
                    break
                except ValueError:
                    continue

        # Revenue patterns - require at least 1 digit
        # Matches: "$391 billion revenue", "revenue of $391B", "annual revenue: $391 billion"
        rev_patterns = [
            r'\$?([\d.,]+)\s*(trillion|billion|million|T|B|M)\s*(?:in\s*)?(?:annual\s*)?revenue',
            r'(?:annual\s*)?revenue[:\s]*(?:of\s*)?\$?([\d.,]+)\s*(trillion|billion|million|T|B|M)\b',
            r'total\s*revenue[:\s]*\$?([\d.,]+)\s*(trillion|billion|million|T|B|M)\b',
        ]
        for pattern in rev_patterns:
            rev_match = re.search(pattern, summary, re.IGNORECASE)
            if rev_match:
                value = rev_match.group(1).replace(",", "")
                try:
                    float(value)
                    unit = rev_match.group(2).upper()
                    unit_map = {"TRILLION": "T", "BILLION": "B", "MILLION": "M"}
                    unit = unit_map.get(unit, unit)
                    metrics["revenue_text"] = f"${value}{unit}"
                    break
                except ValueError:
                    continue

        # P/E ratio - prefer TTM (trailing twelve months) over Forward P/E
        # Order matters: TTM patterns first, then generic P/E (but NOT forward)
        pe_patterns = [
            r'P/?E\s*(?:ratio)?\s*\(?TTM\)?[:\s]*([\d]+\.?[\d]*)',  # P/E (TTM): 37.32
            r'(?:trailing|ttm)\s*P/?E[:\s]*([\d]+\.?[\d]*)',  # Trailing P/E: 37.32
            r'P/?E\s*(?:ratio)?[:\s]*([\d]+\.?[\d]*)',  # Generic P/E (will skip forward due to order)
            r'price[- ]to[- ]earnings[:\s]*([\d]+\.?[\d]*)',
        ]
        for pattern in pe_patterns:
            pe_match = re.search(pattern, summary, re.IGNORECASE)
            if pe_match:
                # Skip if this is actually a "forward P/E" match
                match_start = pe_match.start()
                context_before = summary[max(0, match_start-10):match_start].lower()
                if 'forward' in context_before:
                    continue
                try:
                    pe_value = float(pe_match.group(1))
                    if 0 < pe_value < 10000:  # Sanity check
                        metrics["pe_ratio"] = pe_value
                        break
                except ValueError:
                    continue

        # Employee count - must have at least 2 digits (no single digit counts)
        emp_patterns = [
            r'([\d]{2,3}(?:,[\d]{3})*)\s*(?:\+\s*)?employees',
            r'(?:employs|employing|workforce of)\s*([\d]{2,3}(?:,[\d]{3})*)',
            r'employee\s*count[:\s]*([\d]{2,3}(?:,[\d]{3})*)',
        ]
        for pattern in emp_patterns:
            emp_match = re.search(pattern, summary, re.IGNORECASE)
            if emp_match:
                emp_value = emp_match.group(1).replace(",", "")
                try:
                    emp_num = int(emp_value)
                    if emp_num >= 10:  # Sanity check - at least 10 employees
                        metrics["employees"] = emp_num
                        metrics["employees_text"] = emp_match.group(1)
                        break
                except ValueError:
                    continue

        # Stock price extraction
        price_patterns = [
            r'(?:current\s*)?(?:stock\s*)?price[:\s]*\$([\d]+\.[\d]{2})',
            r'trading\s*at\s*\$([\d]+\.[\d]{2})',
            r'\$([\d]+\.[\d]{2})\s*per\s*share',
        ]
        for pattern in price_patterns:
            price_match = re.search(pattern, summary, re.IGNORECASE)
            if price_match:
                try:
                    price_value = float(price_match.group(1))
                    if 0.01 < price_value < 100000:  # Sanity check
                        metrics["stock_price"] = price_value
                        metrics["stock_price_text"] = f"${price_match.group(1)}"
                        break
                except ValueError:
                    continue

        # EPS extraction
        eps_patterns = [
            r'EPS[:\s]*\$([\d]+\.[\d]+)',
            r'earnings\s*per\s*share[:\s]*\$([\d]+\.[\d]+)',
        ]
        for pattern in eps_patterns:
            eps_match = re.search(pattern, summary, re.IGNORECASE)
            if eps_match:
                try:
                    eps_value = float(eps_match.group(1))
                    metrics["eps"] = eps_value
                    metrics["eps_text"] = f"${eps_match.group(1)}"
                    break
                except ValueError:
                    continue

        return metrics

    def _write_financial_json(self, path: Path):
        """Write financial API data to JSON file."""
        if not self._financial_data:
            return

        data = self._financial_data.to_dict()
        path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")

    def _write_index_file(
        self,
        path: Path,
        company_name: str,
        report_paths: Dict[str, str],
        timestamp: str
    ):
        """Write index/README file for easy navigation."""
        content = f"""# {company_name} Research

**Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M")}

## Report Files

| File | Description |
|------|-------------|
"""
        for file_key, file_config in OUTPUT_FILES.items():
            if file_key in report_paths:
                filename = file_config["filename"]
                description = file_config["description"]
                content += f"| [{filename}]({filename}) | {description} |\n"

        content += """
## Data Files

| File | Description |
|------|-------------|
| [metrics.json](metrics.json) | Machine-readable key metrics |
"""
        if "financial_data" in report_paths:
            content += "| [financial_data.json](financial_data.json) | Raw financial API data |\n"

        content += f"""
## Quick Links

- 📊 [Full Report](00_full_report.md) - Complete comprehensive analysis
- 📈 [Financials](03_financials.md) - Financial performance details
- 🏆 [Competitive Analysis](05_competitive_analysis.md) - Market competition and SWOT
- 📰 [Recent News](06_strategy.md) - Strategy and recent developments

---

*Research generated using Comprehensive Research Runner*
"""
        path.write_text(content, encoding="utf-8")

    def _write_pdf_report(
        self,
        path: Path,
        company_name: str,
        profile: Optional[CompanyProfile],
        analysis: Dict[str, Any]
    ):
        """Write PDF report (requires reportlab or similar)."""
        # Simple text-based PDF fallback
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            c = canvas.Canvas(str(path), pagesize=letter)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, 750, f"{company_name} Research Report")
            c.setFont("Helvetica", 10)
            c.drawString(50, 730, f"Generated: {datetime.now().strftime('%Y-%m-%d')}")

            # Write summary (simplified)
            y = 700
            summary = analysis.get("summary", "")[:2000]
            for line in summary.split('\n')[:50]:
                if y < 50:
                    c.showPage()
                    y = 750
                c.drawString(50, y, line[:100])
                y -= 12

            c.save()
        except ImportError:
            # Fallback: write as text file with .pdf extension
            path.write_text(
                f"{company_name} Research Report\n\n{analysis.get('summary', '')}",
                encoding="utf-8"
            )

    def _write_excel_report(
        self,
        path: Path,
        company_name: str,
        profile: Optional[CompanyProfile],
        analysis: Dict[str, Any]
    ):
        """Write Excel report (requires openpyxl)."""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"

            # Header
            ws['A1'] = "Company Research Report"
            ws['A2'] = company_name
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"

            # Profile info
            if profile:
                ws['A5'] = "Company Profile"
                ws['A6'] = "Industry"
                ws['B6'] = profile.industry
                ws['A7'] = "Country"
                ws['B7'] = profile.country

            # Report sheet
            ws2 = wb.create_sheet("Full Report")
            summary = analysis.get("summary", "")
            for i, line in enumerate(summary.split('\n')[:100], 1):
                ws2[f'A{i}'] = line

            wb.save(path)
        except ImportError:
            print("  [WARNING] openpyxl not installed, skipping Excel")

    def _estimate_cost(self, analysis: Dict[str, Any]) -> float:
        """Estimate cost based on token usage."""
        tokens = analysis.get("tokens", {})
        input_tokens = tokens.get("input", 0)
        output_tokens = tokens.get("output", 0)

        # Haiku pricing
        input_cost = (input_tokens / 1_000_000) * 0.80
        output_cost = (output_tokens / 1_000_000) * 4.00

        # Add estimated search costs
        search_cost = 0.01  # Approximate Tavily cost

        return input_cost + output_cost + search_cost

    def _print_result_summary(self, result: ResearchResult):
        """Print research result summary with completeness details."""
        print(f"\n{'='*70}")
        print(f"[COMPLETE] Research completed for: {result.company_name}")
        print(f"{'='*70}")
        print(f"Duration:      {result.metrics.get('duration_seconds', 0):.1f} seconds")
        print(f"Cost:          ${result.metrics.get('cost_usd', 0):.4f}")
        print(f"Sources:       {result.metrics.get('sources_count', 0)}")
        print(f"Queries:       {result.metrics.get('queries_count', 0)}")
        print(f"Gap-fill iterations: {result.metrics.get('gap_fill_iterations', 0)}")
        print(f"Quality Score: {result.quality_score:.1f}/100")

        # Print completeness summary
        if result.summary:
            section_analysis = self._analyze_section_completeness(result.summary)
            complete = sum(1 for s in section_analysis.values() if s["complete"])
            total = len(section_analysis)
            print(f"\n[COMPLETENESS] {complete}/{total} sections complete")

            # Show incomplete sections
            incomplete = [(name, info) for name, info in section_analysis.items() if not info["complete"]]
            if incomplete:
                print("  Sections needing attention:")
                for section_name, info in incomplete[:5]:
                    display_name = section_name.replace("_", " ").title()
                    score_pct = info["score"] * 100
                    print(f"    - {display_name}: {score_pct:.0f}% complete")

            # Show remaining gaps
            remaining_gaps = result.metrics.get("gap_categories", [])
            if remaining_gaps:
                print(f"\n  Remaining data gaps ({len(remaining_gaps)}):")
                for gap in remaining_gaps[:5]:
                    priority = GAP_PRIORITY.get(gap, 0)
                    priority_label = "HIGH" if priority >= 7 else "MED" if priority >= 5 else "LOW"
                    print(f"    - {gap.replace('_', ' ').title()} [{priority_label}]")

        print(f"\nReports generated:")
        for fmt, path in result.report_paths.items():
            print(f"  - {fmt}: {path}")
        print(f"{'='*70}\n")

    def _print_batch_summary(self, results: List[ResearchResult]):
        """Print batch research summary."""
        successful = [r for r in results if r.success]
        failed = [r for r in results if not r.success]

        print(f"\n{'='*70}")
        print("[BATCH COMPLETE] Market Research Summary")
        print(f"{'='*70}")
        print(f"Total companies: {len(results)}")
        print(f"Successful:      {len(successful)}")
        print(f"Failed:          {len(failed)}")

        if successful:
            total_cost = sum(r.metrics.get('cost_usd', 0) for r in successful)
            total_duration = sum(r.metrics.get('duration_seconds', 0) for r in successful)
            avg_quality = sum(r.quality_score for r in successful) / len(successful)
            print(f"Total cost:      ${total_cost:.4f}")
            print(f"Total duration:  {total_duration:.1f} seconds")
            print(f"Avg quality:     {avg_quality:.1f}/100")

        if failed:
            print(f"\nFailed companies:")
            for r in failed:
                print(f"  - {r.company_name}: {r.error}")

        print(f"{'='*70}\n")

    async def _generate_comparison_report(
        self,
        results: List[ResearchResult],
        market_config: Optional[MarketConfig]
    ):
        """Generate market comparison report."""
        print("\n[COMPARISON] Generating comparison report...")

        market_name = market_config.name if market_config else "Market Comparison"
        safe_name = market_name.lower().replace(" ", "_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        successful_results = [r for r in results if r.success]

        if len(successful_results) < 2:
            print("[WARNING] Not enough successful results for comparison")
            return

        # Create comparison directory
        comparison_dir = self.output_base / "comparisons"
        comparison_dir.mkdir(parents=True, exist_ok=True)

        # Generate comparison content
        content = f"# {market_name} Comparison Report\n\n"
        content += f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
        content += f"**Companies Analyzed:** {len(successful_results)}\n\n"

        content += "## Company Overview\n\n"
        content += "| Company | Industry | Country | Quality Score |\n"
        content += "|---------|----------|---------|---------------|\n"

        for r in successful_results:
            industry = r.profile.industry if r.profile else "N/A"
            country = r.profile.country if r.profile else "N/A"
            content += f"| {r.company_name} | {industry} | {country} | {r.quality_score:.1f} |\n"

        # Write comparison report
        comparison_path = comparison_dir / f"{safe_name}_comparison_{timestamp}.md"
        comparison_path.write_text(content, encoding="utf-8")
        print(f"[COMPARISON] Report saved: {comparison_path}")
